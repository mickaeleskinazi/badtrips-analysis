#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from trip_reports.phenomenological_taxonomy import PHENOMENOLOGICAL_DOMAINS  # noqa: E402
from trip_reports.serious_events import narrative_portion, normalize_space  # noqa: E402


DEFAULT_REPORTS = PROJECT_ROOT / "data" / "processed" / "reports_for_coding.csv"
DEFAULT_CODES = PROJECT_ROOT / "data" / "processed" / "report_codes.csv"
DEFAULT_OUT_CSV = PROJECT_ROOT / "data" / "processed" / "human_review" / "phenomenology_resolution_validation_queue.csv"
DEFAULT_OUT_XLSX = PROJECT_ROOT / "outputs" / "human_review" / "phenomenology_resolution_validation_workbook.xlsx"
DEFAULT_SUMMARY = PROJECT_ROOT / "outputs" / "tables" / "phenomenology_resolution_validation_queue_summary.csv"


STRATA = [
    ("death_acceptance_integration", 80, "critical"),
    ("fear_death_no_resolution", 80, "critical"),
    ("acceptance_surrender", 60, "high"),
    ("integration_positive", 60, "high"),
    ("mystical_religious_entity", 80, "high"),
    ("trauma_autobiographical_insight", 80, "high"),
    ("negative_aftermath", 80, "high"),
    ("interoceptive_loss_control", 80, "high"),
]

YES_NO_UNCLEAR = ["yes", "no", "unclear"]
CENTRAL_CRISIS_OPTIONS = [
    "fear_of_death",
    "fear_of_madness",
    "loss_of_control",
    "interoceptive_threat",
    "paranoia_social_threat",
    "derealization_depersonalization",
    "time_loop_or_eternity",
    "guilt_or_shame",
    "trauma_reexperiencing",
    "mystical_or_religious_crisis",
    "entity_or_god_encounter",
    "social_isolation_or_abandonment",
    "other",
    "unclear",
]
DEATH_THEME_OPTIONS = [
    "none",
    "bodily_death_fear",
    "mortality_realization_self",
    "mortality_of_loved_ones",
    "ego_death_or_self_dissolution",
    "afterlife_god_judgment",
    "actual_death_or_fatality_reported",
    "metaphoric_death_language",
    "unclear",
]
RESOLUTION_STATUS_OPTIONS = [
    "resolved_during_trip",
    "resolved_after_trip",
    "partially_resolved",
    "not_resolved",
    "worsened_or_persisted",
    "unclear",
]
RESOLUTION_MECHANISM_OPTIONS = [
    "acceptance_surrender",
    "social_support_reassurance",
    "change_environment_music",
    "grounding_breathing_body_regulation",
    "prayer_spiritual_frame",
    "medical_or_emergency_intervention",
    "medication_sleep_time_wore_off",
    "purge_vomiting_somatic_release",
    "meaning_making_autobiographical_insight",
    "avoidance_or_behavior_change",
    "other",
    "none_identified",
    "unclear",
]
POST_TRIP_OUTCOME_OPTIONS = [
    "positive_integration",
    "therapeutic_insight",
    "behavior_change_or_abstinence",
    "neutral_recovery",
    "persistent_anxiety_or_derealization",
    "traumatization",
    "psychiatric_deterioration",
    "unclear",
]
CONFIDENCE_OPTIONS = ["high", "medium", "low"]


RESOLUTION_PATTERNS = {
    "social_support_reassurance": [
        r"\bfriend(s)?\b",
        r"\bgirlfriend\b",
        r"\bboyfriend\b",
        r"\bmother\b",
        r"\bfather\b",
        r"\btrip sitter\b",
        r"\btalked me down\b",
        r"\bcalm(ed)? me down\b",
        r"\breassur",
    ],
    "change_environment_music": [
        r"\bchanged? (the )?(room|setting|environment)\b",
        r"\bwent outside\b",
        r"\bwent home\b",
        r"\bturned .* music\b",
        r"\bput on .* music\b",
        r"\bheadphones\b",
    ],
    "grounding_breathing_body_regulation": [
        r"\bbreath(e|ing)\b",
        r"\bground(ed|ing)?\b",
        r"\bmeditat",
        r"\bclosed my eyes\b",
        r"\brelax(ed|ing)?\b",
    ],
    "prayer_spiritual_frame": [
        r"\bpray(ed|ing)?\b",
        r"\bgod\b",
        r"\bjesus\b",
        r"\bangel\b",
        r"\bsoul\b",
        r"\bspiritual\b",
    ],
    "medical_or_emergency_intervention": [
        r"\bhospital\b",
        r"\bambulance\b",
        r"\bdoctor\b",
        r"\bnurse\b",
        r"\bparamedic",
        r"\bemergency room\b",
    ],
    "medication_sleep_time_wore_off": [
        r"\bslept\b",
        r"\bsleep\b",
        r"\bwore off\b",
        r"\bcome down\b",
        r"\bcomedown\b",
        r"\bbenzodiazepine\b",
        r"\bxanax\b",
        r"\bvalium\b",
    ],
    "purge_vomiting_somatic_release": [
        r"\bvomit",
        r"\bpuke",
        r"\bthrowing up\b",
        r"\bthrew up\b",
        r"\bpurge\b",
    ],
    "meaning_making_autobiographical_insight": [
        r"\binsight\b",
        r"\brealized\b",
        r"\blearned\b",
        r"\blesson\b",
        r"\bmeaningful\b",
        r"\btherapy\b",
        r"\bhealing\b",
    ],
    "avoidance_or_behavior_change": [
        r"\bnever .* again\b",
        r"\bstopped using\b",
        r"\bquit\b",
        r"\bavoid(ed)?\b",
        r"\brespect for\b",
    ],
}

DEATH_THEME_PATTERNS = {
    "bodily_death_fear": [
        r"\bthought i (was|would be|might be)? ?dying\b",
        r"\bi (was|am|'m) dying\b",
        r"\bgoing to die\b",
        r"\babout to die\b",
        r"\bheart attack\b",
        r"\boverdose\b",
        r"\bnear[- ]death\b",
    ],
    "mortality_realization_self": [
        r"\bmortality\b",
        r"\bmortal\b",
        r"\bmy own death\b",
        r"\bi will die\b",
        r"\beveryone dies\b",
    ],
    "mortality_of_loved_ones": [
        r"\b(my|her|his) (mother|father|mom|dad|friend|girlfriend|boyfriend|wife|husband).* (die|death|dying)\b",
        r"\b(death|dying) of (my|her|his) (mother|father|mom|dad|friend|girlfriend|boyfriend|wife|husband)\b",
    ],
    "ego_death_or_self_dissolution": [
        r"\bego death\b",
        r"\bego loss\b",
        r"\bi ceased to exist\b",
        r"\bi disappeared\b",
        r"\bno self\b",
    ],
    "afterlife_god_judgment": [
        r"\bgod\b",
        r"\bheaven\b",
        r"\bhell\b",
        r"\bafterlife\b",
        r"\bjudg(e)?ment\b",
        r"\bsoul\b",
    ],
}


def iter_rows(path: Path):
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        yield from csv.DictReader(handle)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_reports(path: Path) -> dict[str, dict[str, str]]:
    return {row["report_id"]: row for row in iter_rows(path)}


def load_codes(path: Path) -> dict[str, dict[str, str]]:
    return {row["report_id"]: row for row in iter_rows(path)}


def compile_domain_patterns() -> dict[str, list[re.Pattern[str]]]:
    compiled = {}
    for domain, config in PHENOMENOLOGICAL_DOMAINS.items():
        compiled[domain] = [re.compile(pattern, re.IGNORECASE) for pattern in config["patterns"]]
    return compiled


def compile_patterns(patterns: dict[str, list[str]]) -> dict[str, list[re.Pattern[str]]]:
    return {name: [re.compile(pattern, re.IGNORECASE) for pattern in values] for name, values in patterns.items()}


def detect_named_patterns(text: str, patterns: dict[str, list[re.Pattern[str]]]) -> dict[str, bool]:
    narrative = narrative_portion(text)
    return {name: any(pattern.search(narrative) for pattern in values) for name, values in patterns.items()}


def snippet_around(text: str, start: int, end: int, radius: int = 220) -> str:
    left = max(0, start - radius)
    right = min(len(text), end + radius)
    prefix = "..." if left else ""
    suffix = "..." if right < len(text) else ""
    return prefix + normalize_space(text[left:right]) + suffix


def collect_snippets(
    text: str,
    pattern_groups: list[tuple[str, dict[str, list[re.Pattern[str]]]]],
    max_snippets: int = 6,
) -> str:
    narrative = narrative_portion(text)
    snippets = []
    seen = set()
    for group_label, patterns in pattern_groups:
        for name, compiled_patterns in patterns.items():
            for pattern in compiled_patterns:
                match = pattern.search(narrative)
                if not match:
                    continue
                snippet = f"{group_label}:{name}: {snippet_around(narrative, match.start(), match.end())}"
                if snippet not in seen:
                    snippets.append(snippet)
                    seen.add(snippet)
                break
            if len(snippets) >= max_snippets:
                return "\n---\n".join(snippets)
    return "\n---\n".join(snippets)


def deterministic_score(*values: str) -> int:
    digest = hashlib.sha256("||".join(values).encode("utf-8")).hexdigest()
    return int(digest[:16], 16)


def code_bool(code_row: dict[str, str], marker: str) -> bool:
    return code_row.get(marker) == "1"


def suggested_crisis_themes(code_row: dict[str, str], domains: dict[str, bool]) -> list[str]:
    themes = []
    for marker, label in [
        ("fear_of_death", "fear_of_death"),
        ("fear_of_madness", "fear_of_madness"),
        ("loss_of_control", "loss_of_control"),
        ("interoceptive_threat", "interoceptive_threat"),
        ("paranoia_social_threat", "paranoia_social_threat"),
        ("derealization_depersonalization", "derealization_depersonalization"),
        ("time_distortion", "time_loop_or_eternity"),
    ]:
        if code_bool(code_row, marker):
            themes.append(label)
    for domain, label in [
        ("religious_feeling", "mystical_or_religious_crisis"),
        ("mystical_experience", "mystical_or_religious_crisis"),
        ("entity_encounter", "entity_or_god_encounter"),
        ("traumatic_reexperiencing", "trauma_reexperiencing"),
    ]:
        if domains.get(domain):
            themes.append(label)
    return sorted(dict.fromkeys(themes))


def suggested_resolution_mechanisms(
    code_row: dict[str, str], resolution_domains: dict[str, bool], resolution_patterns: dict[str, bool]
) -> list[str]:
    mechanisms = []
    if code_bool(code_row, "acceptance_surrender") or resolution_domains.get("acceptance_surrender"):
        mechanisms.append("acceptance_surrender")
    if code_bool(code_row, "social_support"):
        mechanisms.append("social_support_reassurance")
    for name, present in resolution_patterns.items():
        if present:
            mechanisms.append(name)
    return sorted(dict.fromkeys(mechanisms))


def suggested_post_trip_outcomes(domains: dict[str, bool], code_row: dict[str, str]) -> list[str]:
    outcomes = []
    if domains.get("integration_positive_afterwards") or code_bool(code_row, "integration_positive_afterwards"):
        outcomes.append("positive_integration")
    if domains.get("autobiographical_insight"):
        outcomes.append("therapeutic_insight")
    if domains.get("negative_aftermath"):
        outcomes.append("persistent_anxiety_or_derealization")
    return outcomes


def row_score(code_row: dict[str, str], domains: dict[str, bool]) -> int:
    markers = [
        "fear_of_death",
        "acceptance_surrender",
        "integration_positive_afterwards",
        "interoceptive_threat",
        "loss_of_control",
        "fear_of_madness",
        "derealization_depersonalization",
        "social_support",
    ]
    return sum(code_bool(code_row, marker) for marker in markers) + sum(domains.values())


def qualifies(stratum: str, code_row: dict[str, str], domains: dict[str, bool]) -> bool:
    if stratum == "death_acceptance_integration":
        return code_bool(code_row, "fear_of_death") and (
            code_bool(code_row, "acceptance_surrender") or code_bool(code_row, "integration_positive_afterwards")
        )
    if stratum == "fear_death_no_resolution":
        return code_bool(code_row, "fear_of_death") and not code_bool(code_row, "acceptance_surrender") and not code_bool(
            code_row, "integration_positive_afterwards"
        )
    if stratum == "acceptance_surrender":
        return code_bool(code_row, "acceptance_surrender")
    if stratum == "integration_positive":
        return code_bool(code_row, "integration_positive_afterwards")
    if stratum == "mystical_religious_entity":
        return domains.get("mystical_experience", False) or domains.get("religious_feeling", False) or domains.get(
            "entity_encounter", False
        )
    if stratum == "trauma_autobiographical_insight":
        return domains.get("traumatic_reexperiencing", False) or domains.get("autobiographical_insight", False)
    if stratum == "negative_aftermath":
        return domains.get("negative_aftermath", False)
    if stratum == "interoceptive_loss_control":
        return code_bool(code_row, "interoceptive_threat") and code_bool(code_row, "loss_of_control")
    return False


def build_queue(reports_path: Path, codes_path: Path) -> list[dict[str, object]]:
    reports = load_reports(reports_path)
    codes = load_codes(codes_path)
    domain_patterns = compile_domain_patterns()
    death_patterns = compile_patterns(DEATH_THEME_PATTERNS)
    resolution_patterns = compile_patterns(RESOLUTION_PATTERNS)
    detected_domains_by_report: dict[str, dict[str, bool]] = {}
    detected_death_by_report: dict[str, dict[str, bool]] = {}
    detected_resolution_by_report: dict[str, dict[str, bool]] = {}

    for report_id, report in reports.items():
        text = report.get("text", "")
        detected_domains_by_report[report_id] = detect_named_patterns(text, domain_patterns)
        detected_death_by_report[report_id] = detect_named_patterns(text, death_patterns)
        detected_resolution_by_report[report_id] = detect_named_patterns(text, resolution_patterns)

    selected_report_ids: set[str] = set()
    queue_rows: list[dict[str, object]] = []
    for stratum, target_n, priority in STRATA:
        candidates = []
        for report_id, code_row in codes.items():
            if report_id in selected_report_ids:
                continue
            domains = detected_domains_by_report.get(report_id, {})
            if not qualifies(stratum, code_row, domains):
                continue
            candidates.append(
                (
                    -row_score(code_row, domains),
                    deterministic_score(stratum, report_id),
                    report_id,
                )
            )
        candidates.sort()
        for _, __, report_id in candidates[:target_n]:
            selected_report_ids.add(report_id)
            queue_rows.append(
                make_queue_row(
                    stratum=stratum,
                    priority=priority,
                    report=reports[report_id],
                    code_row=codes[report_id],
                    domains=detected_domains_by_report[report_id],
                    death_themes=detected_death_by_report[report_id],
                    resolution_mechanisms=detected_resolution_by_report[report_id],
                    domain_patterns=domain_patterns,
                    death_patterns=death_patterns,
                    resolution_patterns=resolution_patterns,
                )
            )
    return queue_rows


def make_queue_row(
    stratum: str,
    priority: str,
    report: dict[str, str],
    code_row: dict[str, str],
    domains: dict[str, bool],
    death_themes: dict[str, bool],
    resolution_mechanisms: dict[str, bool],
    domain_patterns: dict[str, list[re.Pattern[str]]],
    death_patterns: dict[str, list[re.Pattern[str]]],
    resolution_patterns: dict[str, list[re.Pattern[str]]],
) -> dict[str, object]:
    matched_domains = [name for name, present in domains.items() if present]
    suggested_death = [name for name, present in death_themes.items() if present]
    suggested_resolution = suggested_resolution_mechanisms(code_row, domains, resolution_mechanisms)
    suggested_outcomes = suggested_post_trip_outcomes(domains, code_row)
    return {
        "selection_stratum": stratum,
        "validation_priority": priority,
        "report_id": report.get("report_id", ""),
        "url": report.get("url", ""),
        "title": report.get("title", ""),
        "substance_categories": report.get("substance_categories", ""),
        "erowid_categories": report.get("erowid_categories", ""),
        "target_groups": code_row.get("target_groups", ""),
        "suggested_central_crisis_themes": " | ".join(suggested_crisis_themes(code_row, domains)),
        "suggested_death_theme": " | ".join(suggested_death),
        "suggested_resolution_mechanisms": " | ".join(suggested_resolution),
        "suggested_post_trip_outcome": " | ".join(suggested_outcomes),
        "matched_phenomenological_domains": " | ".join(matched_domains),
        "evidence_snippets": collect_snippets(
            report.get("text", ""),
            [
                ("death", death_patterns),
                ("resolution", resolution_patterns),
                ("phenomenology", domain_patterns),
            ],
        ),
        "central_crisis_theme": "",
        "death_theme": "",
        "mystical_or_religious_content": "",
        "entity_or_god_encounter": "",
        "interoceptive_threat": "",
        "loss_of_control": "",
        "resolution_status": "",
        "primary_resolution_mechanism": "",
        "secondary_resolution_mechanism": "",
        "support_present": "",
        "post_trip_outcome": "",
        "therapeutic_leverage": "",
        "coder_confidence": "",
        "reviewer": "",
        "review_notes": "",
    }


def validation_fieldnames() -> list[str]:
    return [
        "selection_stratum",
        "validation_priority",
        "report_id",
        "url",
        "title",
        "substance_categories",
        "erowid_categories",
        "target_groups",
        "suggested_central_crisis_themes",
        "suggested_death_theme",
        "suggested_resolution_mechanisms",
        "suggested_post_trip_outcome",
        "matched_phenomenological_domains",
        "evidence_snippets",
        "central_crisis_theme",
        "death_theme",
        "mystical_or_religious_content",
        "entity_or_god_encounter",
        "interoceptive_threat",
        "loss_of_control",
        "resolution_status",
        "primary_resolution_mechanism",
        "secondary_resolution_mechanism",
        "support_present",
        "post_trip_outcome",
        "therapeutic_leverage",
        "coder_confidence",
        "reviewer",
        "review_notes",
    ]


def summary_rows(queue_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    counts: Counter[tuple[str, str]] = Counter()
    for row in queue_rows:
        counts[(str(row["selection_stratum"]), str(row["validation_priority"]))] += 1
    return [
        {
            "selection_stratum": stratum,
            "validation_priority": priority,
            "n_rows": count,
        }
        for (stratum, priority), count in sorted(counts.items())
    ]


def add_validation(sheet, col_letter: str, options: list[str], max_row: int) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{max_row}")


def column_letter(fieldnames: list[str], field: str) -> str:
    return get_column_letter(fieldnames.index(field) + 1)


def build_workbook(path: Path, queue_rows: list[dict[str, object]], summary: list[dict[str, object]]) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme_rows = [
        ["Phenomenology / resolution validation", ""],
        ["Purpose", "Code how difficult trips are experienced, whether they resolve, and what appears to help resolution."],
        ["Unit of coding", "One row equals one trip report. Automatic suggestions are only aids."],
        ["Start here", "Code the Validation queue sheet, beginning with critical rows."],
        ["central_crisis_theme", "Primary psychological/phenomenological nucleus of the crisis."],
        ["death_theme", "How death appears: bodily fear, mortality insight, loved ones, ego death, afterlife/God, metaphor, etc."],
        ["resolution_status", "Whether the bad trip resolves during the trip, after the trip, partially, not at all, or worsens."],
        ["resolution_mechanism", "What seems to help: surrender, support, environment/music, grounding, prayer, medical intervention, sleep/time, purge, meaning-making."],
        ["therapeutic_leverage", "Whether the report suggests the difficult experience became clinically/meaningfully useful."],
        ["Important", "This workbook is qualitative validation, not automated classification."],
    ]
    for row_index, row in enumerate(readme_rows, start=1):
        for col_index, value in enumerate(row, start=1):
            readme.cell(row_index, col_index).value = value
    style_readme(readme)

    queue = workbook.create_sheet("Validation queue")
    fields = validation_fieldnames()
    queue.append(fields)
    for row in queue_rows:
        queue.append([row.get(field, "") for field in fields])
    style_table(queue)
    add_validation(queue, column_letter(fields, "central_crisis_theme"), CENTRAL_CRISIS_OPTIONS, queue.max_row)
    add_validation(queue, column_letter(fields, "death_theme"), DEATH_THEME_OPTIONS, queue.max_row)
    for field in [
        "mystical_or_religious_content",
        "entity_or_god_encounter",
        "interoceptive_threat",
        "loss_of_control",
        "support_present",
        "therapeutic_leverage",
    ]:
        add_validation(queue, column_letter(fields, field), YES_NO_UNCLEAR, queue.max_row)
    add_validation(queue, column_letter(fields, "resolution_status"), RESOLUTION_STATUS_OPTIONS, queue.max_row)
    add_validation(queue, column_letter(fields, "primary_resolution_mechanism"), RESOLUTION_MECHANISM_OPTIONS, queue.max_row)
    add_validation(queue, column_letter(fields, "secondary_resolution_mechanism"), RESOLUTION_MECHANISM_OPTIONS, queue.max_row)
    add_validation(queue, column_letter(fields, "post_trip_outcome"), POST_TRIP_OUTCOME_OPTIONS, queue.max_row)
    add_validation(queue, column_letter(fields, "coder_confidence"), CONFIDENCE_OPTIONS, queue.max_row)

    summary_sheet = workbook.create_sheet("Summary")
    summary_fields = ["selection_stratum", "validation_priority", "n_rows"]
    summary_sheet.append(summary_fields)
    for row in summary:
        summary_sheet.append([row[field] for field in summary_fields])
    style_table(summary_sheet)

    lists = workbook.create_sheet("Coding lists")
    list_specs = [
        ("central_crisis_theme", CENTRAL_CRISIS_OPTIONS),
        ("death_theme", DEATH_THEME_OPTIONS),
        ("yes_no_unclear", YES_NO_UNCLEAR),
        ("resolution_status", RESOLUTION_STATUS_OPTIONS),
        ("resolution_mechanism", RESOLUTION_MECHANISM_OPTIONS),
        ("post_trip_outcome", POST_TRIP_OUTCOME_OPTIONS),
        ("coder_confidence", CONFIDENCE_OPTIONS),
    ]
    for col_index, (name, options) in enumerate(list_specs, start=1):
        lists.cell(1, col_index).value = name
        for row_index, option in enumerate(options, start=2):
            lists.cell(row_index, col_index).value = option
    style_table(lists)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def style_readme(sheet) -> None:
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 120
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")


def style_table(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "selection_stratum": 28,
        "validation_priority": 16,
        "report_id": 14,
        "url": 42,
        "title": 42,
        "substance_categories": 36,
        "erowid_categories": 24,
        "target_groups": 32,
        "suggested_central_crisis_themes": 42,
        "suggested_death_theme": 34,
        "suggested_resolution_mechanisms": 42,
        "suggested_post_trip_outcome": 34,
        "matched_phenomenological_domains": 46,
        "evidence_snippets": 100,
        "central_crisis_theme": 28,
        "death_theme": 30,
        "mystical_or_religious_content": 16,
        "entity_or_god_encounter": 16,
        "interoceptive_threat": 16,
        "loss_of_control": 16,
        "resolution_status": 22,
        "primary_resolution_mechanism": 32,
        "secondary_resolution_mechanism": 32,
        "support_present": 16,
        "post_trip_outcome": 28,
        "therapeutic_leverage": 16,
        "coder_confidence": 16,
        "reviewer": 18,
        "review_notes": 56,
    }
    headers = [cell.value for cell in sheet[1]]
    for col_index, header in enumerate(headers, start=1):
        letter = get_column_letter(col_index)
        sheet.column_dimensions[letter].width = widths.get(str(header), 22)
        cell = sheet.cell(1, col_index)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    priority_col = headers.index("validation_priority") + 1 if "validation_priority" in headers else None
    for row in sheet.iter_rows(min_row=2):
        if priority_col:
            priority = row[priority_col - 1].value
            if priority == "critical":
                row[priority_col - 1].fill = PatternFill("solid", fgColor="F4CCCC")
            elif priority == "high":
                row[priority_col - 1].fill = PatternFill("solid", fgColor="FCE5CD")
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a report-level workbook for phenomenology and bad-trip resolution.")
    parser.add_argument("--reports", type=Path, default=DEFAULT_REPORTS)
    parser.add_argument("--codes", type=Path, default=DEFAULT_CODES)
    parser.add_argument("--out-csv", type=Path, default=DEFAULT_OUT_CSV)
    parser.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT_XLSX)
    parser.add_argument("--summary-out", type=Path, default=DEFAULT_SUMMARY)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    queue_rows = build_queue(args.reports, args.codes)
    summary = summary_rows(queue_rows)
    fields = validation_fieldnames()
    write_csv(args.out_csv, fields, queue_rows)
    write_csv(args.summary_out, ["selection_stratum", "validation_priority", "n_rows"], summary)
    build_workbook(args.out_xlsx, queue_rows, summary)

    print(f"Validation rows: {len(queue_rows)}")
    print(f"Strata sampled: {len(summary)}")
    print(f"CSV written to: {args.out_csv}")
    print(f"Workbook written to: {args.out_xlsx}")
    print(f"Summary written to: {args.summary_out}")
    for row in summary:
        print(f"  {row['validation_priority']} | {row['selection_stratum']}: {row['n_rows']}")


if __name__ == "__main__":
    main()
