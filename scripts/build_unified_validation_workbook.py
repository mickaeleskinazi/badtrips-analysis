#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]

DEFAULT_AE_LEGAL_QUEUE = (
    PROJECT_ROOT / "data" / "processed" / "human_review" / "broad_category_report_validation_queue.csv"
)
DEFAULT_PHENO_QUEUE = (
    PROJECT_ROOT / "data" / "processed" / "human_review" / "phenomenology_resolution_validation_queue.csv"
)
DEFAULT_OUT_CSV = PROJECT_ROOT / "data" / "processed" / "human_review" / "unified_report_validation_queue.csv"
DEFAULT_OUT_XLSX = PROJECT_ROOT / "outputs" / "human_review" / "unified_report_validation_workbook.xlsx"
DEFAULT_SUMMARY = PROJECT_ROOT / "outputs" / "tables" / "unified_report_validation_queue_summary.csv"


PRIORITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "": 4}

REVIEW_STATUS_OPTIONS = ["not_started", "in_progress", "complete", "exclude", "unclear"]
YES_NO_UNCLEAR_NA = ["yes", "no", "unclear", "not_applicable"]
AE_FALSE_POSITIVE_OPTIONS = [
    "none",
    "metaphor_or_idiom",
    "hypothetical_or_fear_only",
    "historical_or_baseline",
    "third_party_not_author",
    "boilerplate_or_editorial",
    "wrong_meaning",
    "insufficient_context",
    "not_applicable",
]
SEVERITY_OPTIONS = ["none", "mild", "moderate", "severe", "life_threatening", "death", "unclear"]
SUBSTANCE_ROLE_OPTIONS = ["probable", "possible", "unlikely", "unclear", "not_applicable"]
LEGAL_OUTCOME_OPTIONS = [
    "none",
    "police_contact",
    "arrest_or_custody",
    "charge_or_court",
    "prison",
    "unclear",
    "not_applicable",
]
PSYCHIATRIC_CONTEXT_OPTIONS = [
    "none_reported",
    "baseline_psychiatric_history",
    "acute_decompensation",
    "mania_or_bipolar",
    "depression",
    "psychosis_or_delirium",
    "suicidality",
    "unclear",
]
CENTRAL_CRISIS_OPTIONS = [
    "fear_of_death",
    "fear_of_madness",
    "loss_of_control",
    "interoceptive_threat",
    "paranoia_social_threat",
    "derealization_depersonalization",
    "time_loop_or_eternity",
    "guilt_or_shame",
    "trauma_reexperiencing",
    "mystical_or_religious_crisis",
    "entity_or_god_encounter",
    "social_isolation_or_abandonment",
    "other",
    "unclear",
    "not_applicable",
]
DEATH_THEME_OPTIONS = [
    "none",
    "bodily_death_fear",
    "mortality_realization_self",
    "mortality_of_loved_ones",
    "ego_death_or_self_dissolution",
    "afterlife_god_judgment",
    "actual_death_or_fatality_reported",
    "metaphoric_death_language",
    "unclear",
    "not_applicable",
]
RESOLUTION_STATUS_OPTIONS = [
    "resolved_during_trip",
    "resolved_after_trip",
    "partially_resolved",
    "not_resolved",
    "worsened_or_persisted",
    "unclear",
    "not_applicable",
]
RESOLUTION_MECHANISM_OPTIONS = [
    "acceptance_surrender",
    "social_support_reassurance",
    "change_environment_music",
    "grounding_breathing_body_regulation",
    "prayer_spiritual_frame",
    "medical_or_emergency_intervention",
    "medication_sleep_time_wore_off",
    "purge_vomiting_somatic_release",
    "meaning_making_autobiographical_insight",
    "avoidance_or_behavior_change",
    "other",
    "none_identified",
    "unclear",
    "not_applicable",
]
POST_TRIP_OUTCOME_OPTIONS = [
    "positive_integration",
    "therapeutic_insight",
    "behavior_change_or_abstinence",
    "neutral_recovery",
    "persistent_anxiety_or_derealization",
    "traumatization",
    "psychiatric_deterioration",
    "unclear",
    "not_applicable",
]
CONFIDENCE_OPTIONS = ["high", "medium", "low"]


def iter_rows(path: Path):
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        yield from csv.DictReader(handle)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def split_pipe(value: str) -> list[str]:
    return [item.strip() for item in (value or "").split(" | ") if item.strip()]


def unique_join(values: list[str], sep: str = " | ") -> str:
    return sep.join(dict.fromkeys(item for item in values if item))


def priority_rank(priority: str) -> int:
    return PRIORITY_ORDER.get(priority or "", 4)


def strongest_priority(priorities: list[str]) -> str:
    if not priorities:
        return ""
    return sorted(priorities, key=priority_rank)[0]


def truncate_text(text: str, max_chars: int = 4200) -> str:
    text = text or ""
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rstrip() + "\n...[truncated]"


def combine_snippets(rows: list[dict[str, str]], field: str, max_blocks: int = 8, max_chars: int = 4200) -> str:
    blocks = []
    seen = set()
    for row in rows:
        label = row.get("broad_cohort") or row.get("selection_stratum") or ""
        snippets = row.get(field, "")
        if not snippets:
            continue
        block = f"{label}\n{snippets}" if label else snippets
        if block not in seen:
            blocks.append(block)
            seen.add(block)
        if len(blocks) >= max_blocks:
            break
    return truncate_text("\n\n====\n\n".join(blocks), max_chars=max_chars)


def aggregate_ae_legal(rows: list[dict[str, str]]) -> dict[str, object]:
    if not rows:
        return {
            "ae_legal_candidate_categories": "",
            "ae_legal_candidate_terms": "",
            "ae_legal_candidate_priorities": "",
            "ae_legal_evidence_snippets": "",
            "ae_legal_source_rows": 0,
        }
    categories = [row.get("broad_cohort", "") for row in rows]
    terms_by_category = []
    for row in rows:
        category = row.get("broad_cohort", "")
        terms = row.get("matched_terms", "")
        if category and terms:
            terms_by_category.append(f"{category}: {terms}")
    return {
        "ae_legal_candidate_categories": unique_join(categories),
        "ae_legal_candidate_terms": unique_join(terms_by_category, sep="\n"),
        "ae_legal_candidate_priorities": unique_join([row.get("validation_priority", "") for row in rows]),
        "ae_legal_evidence_snippets": combine_snippets(rows, "evidence_snippets"),
        "ae_legal_source_rows": len(rows),
    }


def aggregate_pheno(rows: list[dict[str, str]]) -> dict[str, object]:
    if not rows:
        return {
            "pheno_selection_strata": "",
            "suggested_central_crisis_themes": "",
            "suggested_death_theme": "",
            "suggested_resolution_mechanisms": "",
            "suggested_post_trip_outcome": "",
            "matched_phenomenological_domains": "",
            "pheno_evidence_snippets": "",
            "pheno_source_rows": 0,
        }
    return {
        "pheno_selection_strata": unique_join([row.get("selection_stratum", "") for row in rows]),
        "suggested_central_crisis_themes": unique_join(
            sum((split_pipe(row.get("suggested_central_crisis_themes", "")) for row in rows), [])
        ),
        "suggested_death_theme": unique_join(
            sum((split_pipe(row.get("suggested_death_theme", "")) for row in rows), [])
        ),
        "suggested_resolution_mechanisms": unique_join(
            sum((split_pipe(row.get("suggested_resolution_mechanisms", "")) for row in rows), [])
        ),
        "suggested_post_trip_outcome": unique_join(
            sum((split_pipe(row.get("suggested_post_trip_outcome", "")) for row in rows), [])
        ),
        "matched_phenomenological_domains": unique_join(
            sum((split_pipe(row.get("matched_phenomenological_domains", "")) for row in rows), [])
        ),
        "pheno_evidence_snippets": combine_snippets(rows, "evidence_snippets", max_blocks=4),
        "pheno_source_rows": len(rows),
    }


def first_metadata(report_id: str, ae_rows: list[dict[str, str]], pheno_rows: list[dict[str, str]]) -> dict[str, str]:
    for row in ae_rows + pheno_rows:
        if row:
            return {
                "report_id": report_id,
                "url": row.get("url", ""),
                "title": row.get("title", ""),
                "substance_categories": row.get("substance_categories", ""),
                "erowid_categories": row.get("erowid_categories", ""),
                "target_groups": row.get("target_groups", ""),
            }
    return {
        "report_id": report_id,
        "url": "",
        "title": "",
        "substance_categories": "",
        "erowid_categories": "",
        "target_groups": "",
    }


def build_unified_rows(ae_legal_path: Path, pheno_path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    ae_rows_by_report: dict[str, list[dict[str, str]]] = defaultdict(list)
    pheno_rows_by_report: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in iter_rows(ae_legal_path):
        ae_rows_by_report[row["report_id"]].append(row)
    for row in iter_rows(pheno_path):
        pheno_rows_by_report[row["report_id"]].append(row)

    report_ids = sorted(set(ae_rows_by_report) | set(pheno_rows_by_report))
    unified_rows: list[dict[str, object]] = []
    category_detail_rows: list[dict[str, object]] = []

    for report_id in report_ids:
        ae_rows = ae_rows_by_report.get(report_id, [])
        pheno_rows = pheno_rows_by_report.get(report_id, [])
        source_queue = "both" if ae_rows and pheno_rows else "ae_legal" if ae_rows else "phenomenology"
        priorities = [row.get("validation_priority", "") for row in ae_rows + pheno_rows]
        metadata = first_metadata(report_id, ae_rows, pheno_rows)
        ae = aggregate_ae_legal(ae_rows)
        pheno = aggregate_pheno(pheno_rows)
        unified_rows.append(
            {
                "source_queue": source_queue,
                "validation_priority": strongest_priority(priorities),
                **metadata,
                **ae,
                **pheno,
                "report_review_status": "",
                "ae_legal_event_present": "",
                "ae_legal_confirmed_categories": "",
                "ae_legal_false_positive_type": "",
                "ae_legal_severity": "",
                "ae_legal_substance_role": "",
                "ae_legal_legal_outcome": "",
                "ae_legal_psychiatric_context": "",
                "central_crisis_theme": "",
                "death_theme": "",
                "mystical_or_religious_content": "",
                "entity_or_god_encounter": "",
                "interoceptive_threat": "",
                "loss_of_control": "",
                "resolution_status": "",
                "primary_resolution_mechanism": "",
                "secondary_resolution_mechanism": "",
                "support_present": "",
                "post_trip_outcome": "",
                "therapeutic_leverage": "",
                "coder_confidence": "",
                "reviewer": "",
                "review_notes": "",
            }
        )
        for row in ae_rows:
            category_detail_rows.append(
                {
                    "report_id": report_id,
                    "source": "ae_legal",
                    "category_or_stratum": row.get("broad_cohort", ""),
                    "priority": row.get("validation_priority", ""),
                    "matched_terms": row.get("matched_terms", ""),
                    "selection_reason": row.get("selection_reason", ""),
                }
            )
        for row in pheno_rows:
            category_detail_rows.append(
                {
                    "report_id": report_id,
                    "source": "phenomenology",
                    "category_or_stratum": row.get("selection_stratum", ""),
                    "priority": row.get("validation_priority", ""),
                    "matched_terms": row.get("suggested_central_crisis_themes", ""),
                    "selection_reason": "",
                }
            )

    unified_rows.sort(
        key=lambda row: (
            priority_rank(str(row["validation_priority"])),
            {"both": 0, "ae_legal": 1, "phenomenology": 2}.get(str(row["source_queue"]), 3),
            str(row["report_id"]),
        )
    )
    return unified_rows, category_detail_rows


def unified_fieldnames() -> list[str]:
    return [
        "source_queue",
        "validation_priority",
        "report_id",
        "url",
        "title",
        "substance_categories",
        "erowid_categories",
        "target_groups",
        "ae_legal_candidate_categories",
        "ae_legal_candidate_terms",
        "ae_legal_candidate_priorities",
        "ae_legal_evidence_snippets",
        "ae_legal_source_rows",
        "pheno_selection_strata",
        "suggested_central_crisis_themes",
        "suggested_death_theme",
        "suggested_resolution_mechanisms",
        "suggested_post_trip_outcome",
        "matched_phenomenological_domains",
        "pheno_evidence_snippets",
        "pheno_source_rows",
        "report_review_status",
        "ae_legal_event_present",
        "ae_legal_confirmed_categories",
        "ae_legal_false_positive_type",
        "ae_legal_severity",
        "ae_legal_substance_role",
        "ae_legal_legal_outcome",
        "ae_legal_psychiatric_context",
        "central_crisis_theme",
        "death_theme",
        "mystical_or_religious_content",
        "entity_or_god_encounter",
        "interoceptive_threat",
        "loss_of_control",
        "resolution_status",
        "primary_resolution_mechanism",
        "secondary_resolution_mechanism",
        "support_present",
        "post_trip_outcome",
        "therapeutic_leverage",
        "coder_confidence",
        "reviewer",
        "review_notes",
    ]


def build_summary_rows(unified_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    counts = Counter((str(row["source_queue"]), str(row["validation_priority"])) for row in unified_rows)
    return [
        {
            "source_queue": source,
            "validation_priority": priority,
            "n_reports": count,
        }
        for (source, priority), count in sorted(counts.items(), key=lambda item: (priority_rank(item[0][1]), item[0][0]))
    ]


def add_validation(sheet, col_letter: str, options: list[str], max_row: int) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{max_row}")


def column_letter(fieldnames: list[str], field: str) -> str:
    return get_column_letter(fieldnames.index(field) + 1)


def build_workbook(
    path: Path,
    unified_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    summary_rows: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme_rows = [
        ["Unified report validation", ""],
        ["Purpose", "Read each trip report once and code both AE/forensic/legal outcomes and phenomenology/resolution."],
        ["Main sheet", "Use Unified validation queue. One row equals one unique report."],
        ["AE/legal coding", "Use ae_legal_* columns when the report has candidate severe AE or forensic/legal content."],
        ["Phenomenology coding", "Use central_crisis_theme, death_theme, resolution_status, mechanisms, and post_trip_outcome."],
        ["Detail sheet", "Category detail shows which source categories/strata generated each report row."],
        ["Important", "Automatic suggestions and snippets are screening aids, not final labels."],
    ]
    for row_idx, row in enumerate(readme_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            readme.cell(row_idx, col_idx).value = value
    style_readme(readme)

    fields = unified_fieldnames()
    queue = workbook.create_sheet("Unified validation queue")
    queue.append(fields)
    for row in unified_rows:
        queue.append([row.get(field, "") for field in fields])
    style_table(queue)
    add_main_validations(queue, fields)

    summary = workbook.create_sheet("Summary")
    summary_fields = ["source_queue", "validation_priority", "n_reports"]
    summary.append(summary_fields)
    for row in summary_rows:
        summary.append([row[field] for field in summary_fields])
    style_table(summary)

    details = workbook.create_sheet("Category detail")
    detail_fields = ["report_id", "source", "category_or_stratum", "priority", "matched_terms", "selection_reason"]
    details.append(detail_fields)
    for row in detail_rows:
        details.append([row[field] for field in detail_fields])
    style_table(details)

    lists = workbook.create_sheet("Coding lists")
    specs = [
        ("report_review_status", REVIEW_STATUS_OPTIONS),
        ("yes_no_unclear_na", YES_NO_UNCLEAR_NA),
        ("ae_false_positive", AE_FALSE_POSITIVE_OPTIONS),
        ("severity", SEVERITY_OPTIONS),
        ("substance_role", SUBSTANCE_ROLE_OPTIONS),
        ("legal_outcome", LEGAL_OUTCOME_OPTIONS),
        ("psychiatric_context", PSYCHIATRIC_CONTEXT_OPTIONS),
        ("central_crisis", CENTRAL_CRISIS_OPTIONS),
        ("death_theme", DEATH_THEME_OPTIONS),
        ("resolution_status", RESOLUTION_STATUS_OPTIONS),
        ("resolution_mechanism", RESOLUTION_MECHANISM_OPTIONS),
        ("post_trip_outcome", POST_TRIP_OUTCOME_OPTIONS),
        ("confidence", CONFIDENCE_OPTIONS),
    ]
    for col_idx, (name, options) in enumerate(specs, start=1):
        lists.cell(1, col_idx).value = name
        for row_idx, option in enumerate(options, start=2):
            lists.cell(row_idx, col_idx).value = option
    style_table(lists)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def add_main_validations(sheet, fields: list[str]) -> None:
    max_row = sheet.max_row
    add_validation(sheet, column_letter(fields, "report_review_status"), REVIEW_STATUS_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "ae_legal_event_present"), YES_NO_UNCLEAR_NA, max_row)
    add_validation(sheet, column_letter(fields, "ae_legal_false_positive_type"), AE_FALSE_POSITIVE_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "ae_legal_severity"), SEVERITY_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "ae_legal_substance_role"), SUBSTANCE_ROLE_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "ae_legal_legal_outcome"), LEGAL_OUTCOME_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "ae_legal_psychiatric_context"), PSYCHIATRIC_CONTEXT_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "central_crisis_theme"), CENTRAL_CRISIS_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "death_theme"), DEATH_THEME_OPTIONS, max_row)
    for field in [
        "mystical_or_religious_content",
        "entity_or_god_encounter",
        "interoceptive_threat",
        "loss_of_control",
        "support_present",
        "therapeutic_leverage",
    ]:
        add_validation(sheet, column_letter(fields, field), YES_NO_UNCLEAR_NA, max_row)
    add_validation(sheet, column_letter(fields, "resolution_status"), RESOLUTION_STATUS_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "primary_resolution_mechanism"), RESOLUTION_MECHANISM_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "secondary_resolution_mechanism"), RESOLUTION_MECHANISM_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "post_trip_outcome"), POST_TRIP_OUTCOME_OPTIONS, max_row)
    add_validation(sheet, column_letter(fields, "coder_confidence"), CONFIDENCE_OPTIONS, max_row)


def style_readme(sheet) -> None:
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 120
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")


def style_table(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "source_queue": 14,
        "validation_priority": 15,
        "report_id": 13,
        "url": 36,
        "title": 38,
        "substance_categories": 32,
        "erowid_categories": 22,
        "target_groups": 28,
        "ae_legal_candidate_categories": 34,
        "ae_legal_candidate_terms": 46,
        "ae_legal_candidate_priorities": 18,
        "ae_legal_evidence_snippets": 80,
        "ae_legal_source_rows": 10,
        "pheno_selection_strata": 34,
        "suggested_central_crisis_themes": 42,
        "suggested_death_theme": 34,
        "suggested_resolution_mechanisms": 42,
        "suggested_post_trip_outcome": 34,
        "matched_phenomenological_domains": 42,
        "pheno_evidence_snippets": 80,
        "pheno_source_rows": 10,
        "report_review_status": 18,
        "ae_legal_event_present": 18,
        "ae_legal_confirmed_categories": 34,
        "ae_legal_false_positive_type": 26,
        "ae_legal_severity": 18,
        "ae_legal_substance_role": 18,
        "ae_legal_legal_outcome": 22,
        "ae_legal_psychiatric_context": 28,
        "central_crisis_theme": 28,
        "death_theme": 28,
        "mystical_or_religious_content": 16,
        "entity_or_god_encounter": 16,
        "interoceptive_threat": 16,
        "loss_of_control": 16,
        "resolution_status": 22,
        "primary_resolution_mechanism": 30,
        "secondary_resolution_mechanism": 30,
        "support_present": 16,
        "post_trip_outcome": 28,
        "therapeutic_leverage": 16,
        "coder_confidence": 16,
        "reviewer": 18,
        "review_notes": 52,
        "matched_terms": 42,
        "selection_reason": 32,
    }
    headers = [cell.value for cell in sheet[1]]
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        sheet.column_dimensions[letter].width = widths.get(str(header), 22)
        cell = sheet.cell(1, col_idx)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    priority_col = headers.index("validation_priority") + 1 if "validation_priority" in headers else None
    for row in sheet.iter_rows(min_row=2):
        if priority_col:
            priority = row[priority_col - 1].value
            if priority == "critical":
                row[priority_col - 1].fill = PatternFill("solid", fgColor="F4CCCC")
            elif priority == "high":
                row[priority_col - 1].fill = PatternFill("solid", fgColor="FCE5CD")
            elif priority == "medium":
                row[priority_col - 1].fill = PatternFill("solid", fgColor="FFF2CC")
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build one report-level validation workbook for AE/legal and phenomenology.")
    parser.add_argument("--ae-legal-queue", type=Path, default=DEFAULT_AE_LEGAL_QUEUE)
    parser.add_argument("--pheno-queue", type=Path, default=DEFAULT_PHENO_QUEUE)
    parser.add_argument("--out-csv", type=Path, default=DEFAULT_OUT_CSV)
    parser.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT_XLSX)
    parser.add_argument("--summary-out", type=Path, default=DEFAULT_SUMMARY)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    unified_rows, detail_rows = build_unified_rows(args.ae_legal_queue, args.pheno_queue)
    summary_rows = build_summary_rows(unified_rows)
    write_csv(args.out_csv, unified_fieldnames(), unified_rows)
    write_csv(args.summary_out, ["source_queue", "validation_priority", "n_reports"], summary_rows)
    build_workbook(args.out_xlsx, unified_rows, detail_rows, summary_rows)

    print(f"Unified reports: {len(unified_rows)}")
    print(f"Category/detail rows: {len(detail_rows)}")
    print(f"CSV written to: {args.out_csv}")
    print(f"Workbook written to: {args.out_xlsx}")
    print(f"Summary written to: {args.summary_out}")
    for row in summary_rows:
        print(f"  {row['validation_priority']} | {row['source_queue']}: {row['n_reports']}")


if __name__ == "__main__":
    main()
