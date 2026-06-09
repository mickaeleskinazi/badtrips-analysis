#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from trip_reports.serious_events import narrative_portion, normalize_space  # noqa: E402


DEFAULT_REPORTS = PROJECT_ROOT / "data" / "processed" / "reports_for_coding.csv"
DEFAULT_FLAGS = PROJECT_ROOT / "data" / "processed" / "broad_category_report_flags.csv"
DEFAULT_OUT_CSV = PROJECT_ROOT / "data" / "processed" / "human_review" / "broad_category_report_validation_queue.csv"
DEFAULT_OUT_XLSX = PROJECT_ROOT / "outputs" / "human_review" / "broad_category_report_validation_workbook.xlsx"
DEFAULT_SUMMARY = PROJECT_ROOT / "outputs" / "tables" / "broad_category_validation_queue_summary.csv"


VALIDATION_CATEGORIES = {
    "Arrest / detention / prison": {"target_n": 45, "priority": "high"},
    "Charges / court / probation": {"target_n": 45, "priority": "high"},
    "Criminalized behavior / offences": {"target_n": 45, "priority": "high"},
    "Police contact / law enforcement": {"target_n": 45, "priority": "high"},
    "Violence / assault / weapons": {"target_n": 60, "priority": "high"},
    "Homicide / death investigation": {"target_n": 75, "priority": "critical"},
    "Sexual violence / exploitation": {"target_n": 75, "priority": "critical"},
    "Accidents / traffic / public safety": {"target_n": 60, "priority": "high"},
    "Forensic psychiatry interface": {"target_n": 50, "priority": "high"},
    "Psychiatric decompensation": {"target_n": 60, "priority": "high"},
    "Behavioral danger / disorganization": {"target_n": 45, "priority": "high"},
    "Suicide / self-harm": {"target_n": 75, "priority": "critical"},
    "Hospital / emergency care": {"target_n": 45, "priority": "medium"},
    "Emergency rescue / ambulance": {"target_n": 45, "priority": "medium"},
    "Somatic / life-threatening medical AE": {"target_n": 60, "priority": "high"},
    "Death / fatality reported": {"target_n": 75, "priority": "critical"},
    "Accident / trauma / fall from height": {"target_n": 75, "priority": "critical"},
    "Accident / trauma": {"target_n": 45, "priority": "high"},
}

SKIP_BROAD_CATEGORIES = {
    "Any forensic/legal relevance",
    "Any severe adverse event",
    "High-confidence severe adverse event",
    "Justice-system involvement",
    "Police / legal involvement",
}

EVENT_PRESENT_OPTIONS = ["yes", "no", "unclear"]
FALSE_POSITIVE_OPTIONS = [
    "none",
    "metaphor_or_idiom",
    "hypothetical_or_fear_only",
    "historical_or_baseline",
    "third_party_not_author",
    "boilerplate_or_editorial",
    "wrong_meaning",
    "insufficient_context",
]
SEVERITY_OPTIONS = ["none", "mild", "moderate", "severe", "life_threatening", "death", "unclear"]
SUBSTANCE_ROLE_OPTIONS = ["probable", "possible", "unlikely", "unclear"]
LEGAL_OUTCOME_OPTIONS = ["none", "police_contact", "arrest_or_custody", "charge_or_court", "prison", "unclear"]
PSYCHIATRIC_CONTEXT_OPTIONS = [
    "none_reported",
    "baseline_psychiatric_history",
    "acute_decompensation",
    "mania_or_bipolar",
    "depression",
    "psychosis_or_delirium",
    "suicidality",
    "unclear",
]


def iter_rows(path: Path):
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        yield from csv.DictReader(handle)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def deterministic_score(*values: str) -> int:
    digest = hashlib.sha256("||".join(values).encode("utf-8")).hexdigest()
    return int(digest[:16], 16)


def load_reports(path: Path) -> dict[str, dict[str, str]]:
    return {row["report_id"]: row for row in iter_rows(path)}


def broad_category_columns(fieldnames: list[str]) -> dict[str, dict[str, str]]:
    columns: dict[str, dict[str, str]] = {}
    for field in fieldnames:
        if not field.startswith("terms__"):
            continue
        slug = field.removeprefix("terms__")
        columns[slug] = {
            "flag": f"flag__{slug}",
            "n_terms": f"n_terms__{slug}",
            "terms": field,
        }
    return columns


def split_terms(value: str) -> list[str]:
    return [term.strip() for term in (value or "").split(" | ") if term.strip()]


def word_pattern(term: str) -> re.Pattern[str]:
    parts = [re.escape(part) for part in term.split()]
    pattern = r"\b" + r"\s+".join(parts) + r"\b"
    return re.compile(pattern, flags=re.IGNORECASE)


def snippet_around(text: str, start: int, end: int, radius: int = 220) -> str:
    left = max(0, start - radius)
    right = min(len(text), end + radius)
    prefix = "..." if left else ""
    suffix = "..." if right < len(text) else ""
    return prefix + normalize_space(text[left:right]) + suffix


def evidence_snippets(text: str, terms: list[str], max_snippets: int = 4) -> str:
    narrative = narrative_portion(text)
    snippets = []
    seen = set()
    for term in terms:
        for match in word_pattern(term).finditer(narrative):
            snippet = f"{term}: {snippet_around(narrative, match.start(), match.end())}"
            if snippet not in seen:
                snippets.append(snippet)
                seen.add(snippet)
            break
        if len(snippets) >= max_snippets:
            break
    return "\n---\n".join(snippets)


def validation_categories_from_flags(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
    slug_columns = broad_category_columns(fieldnames)
    return {
        broad_name: slug_columns[slugify(broad_name)]
        for broad_name in VALIDATION_CATEGORIES
        if slugify(broad_name) in slug_columns and broad_name not in SKIP_BROAD_CATEGORIES
    }


def build_queue(reports_path: Path, flags_path: Path) -> list[dict[str, object]]:
    reports = load_reports(reports_path)
    category_columns = validation_categories_from_flags(flags_path)
    candidates_by_category: dict[str, list[dict[str, object]]] = defaultdict(list)

    for flag_row in iter_rows(flags_path):
        report_id = flag_row["report_id"]
        report = reports.get(report_id, {})
        for broad_name, columns in category_columns.items():
            if flag_row.get(columns["flag"]) != "1":
                continue
            terms = split_terms(flag_row.get(columns["terms"], ""))
            if not terms:
                continue
            category_meta = VALIDATION_CATEGORIES[broad_name]
            candidates_by_category[broad_name].append(
                {
                    "validation_priority": category_meta["priority"],
                    "analysis_domain": domain_for_category(broad_name),
                    "broad_cohort": broad_name,
                    "report_id": report_id,
                    "url": flag_row.get("url", ""),
                    "title": flag_row.get("title", ""),
                    "substance_categories": flag_row.get("substance_categories", ""),
                    "erowid_categories": flag_row.get("erowid_categories", ""),
                    "matched_terms": " | ".join(terms),
                    "n_matched_terms_in_category": len(terms),
                    "evidence_snippets": evidence_snippets(report.get("text", ""), terms),
                    "selection_reason": "",
                    "event_present": "",
                    "false_positive_type": "",
                    "severity": "",
                    "substance_role": "",
                    "legal_outcome": "",
                    "psychiatric_context": "",
                    "reviewer": "",
                    "review_notes": "",
                }
            )

    selected: list[dict[str, object]] = []
    for broad_name, candidates in sorted(candidates_by_category.items()):
        target_n = VALIDATION_CATEGORIES[broad_name]["target_n"]
        candidates.sort(
            key=lambda row: (
                -int(row["n_matched_terms_in_category"]),
                deterministic_score(broad_name, row["report_id"]),
            )
        )
        for row in candidates[:target_n]:
            row["selection_reason"] = f"top_{target_n}_by_term_count_then_deterministic_sample"
            selected.append(row)

    selected.sort(
        key=lambda row: (
            priority_rank(str(row["validation_priority"])),
            str(row["analysis_domain"]),
            str(row["broad_cohort"]),
            str(row["report_id"]),
        )
    )
    return selected


def domain_for_category(broad_name: str) -> str:
    forensic = {
        "Arrest / detention / prison",
        "Charges / court / probation",
        "Criminalized behavior / offences",
        "Police contact / law enforcement",
        "Violence / assault / weapons",
        "Homicide / death investigation",
        "Sexual violence / exploitation",
        "Accidents / traffic / public safety",
    }
    severe = {
        "Psychiatric decompensation",
        "Behavioral danger / disorganization",
        "Hospital / emergency care",
        "Emergency rescue / ambulance",
        "Somatic / life-threatening medical AE",
        "Death / fatality reported",
    }
    shared = {
        "Forensic psychiatry interface",
        "Suicide / self-harm",
        "Accident / trauma / fall from height",
        "Accident / trauma",
    }
    if broad_name in forensic:
        return "Forensic/legal"
    if broad_name in severe:
        return "Severe AE"
    if broad_name in shared:
        return "Forensic/legal + Severe AE"
    return "Other"


def priority_rank(priority: str) -> int:
    return {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(priority, 4)


def summary_rows(queue_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    counts: Counter[tuple[str, str, str]] = Counter()
    for row in queue_rows:
        counts[
            (
                str(row["analysis_domain"]),
                str(row["broad_cohort"]),
                str(row["validation_priority"]),
            )
        ] += 1
    return [
        {
            "analysis_domain": domain,
            "broad_cohort": broad,
            "validation_priority": priority,
            "n_validation_rows": count,
        }
        for (domain, broad, priority), count in sorted(
            counts.items(), key=lambda item: (priority_rank(item[0][2]), item[0][0], item[0][1])
        )
    ]


def add_validation(sheet, col_letter: str, options: list[str], max_row: int) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{max_row}")


def build_workbook(path: Path, queue_rows: list[dict[str, object]], summary: list[dict[str, object]]) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme_rows = [
        ["Broad-category report validation", ""],
        ["Purpose", "Validate whether screened trip reports contain true severe adverse events or forensic/legal events."],
        ["How to start", "Code the Validation queue sheet. Focus first on critical rows."],
        ["event_present", "yes/no/unclear: whether the event actually occurred in the report."],
        ["false_positive_type", "Use only when event_present is no or unclear."],
        ["severity", "Clinical/legal severity of the event described."],
        ["substance_role", "Your judgment of whether the substance plausibly contributed."],
        ["legal_outcome", "Code only concrete legal outcomes, not fear of legal trouble."],
        ["psychiatric_context", "Baseline psychiatric history or acute psychiatric decompensation."],
        ["Important", "This is a validation queue, not a prevalence table."],
    ]
    for row_index, row in enumerate(readme_rows, start=1):
        for col_index, value in enumerate(row, start=1):
            readme.cell(row_index, col_index).value = value
    style_readme(readme)

    queue_sheet = workbook.create_sheet("Validation queue")
    fieldnames = validation_fieldnames()
    queue_sheet.append(fieldnames)
    for row in queue_rows:
        queue_sheet.append([row.get(field, "") for field in fieldnames])
    style_table(queue_sheet)
    add_validation(queue_sheet, column_letter(fieldnames, "event_present"), EVENT_PRESENT_OPTIONS, queue_sheet.max_row)
    add_validation(queue_sheet, column_letter(fieldnames, "false_positive_type"), FALSE_POSITIVE_OPTIONS, queue_sheet.max_row)
    add_validation(queue_sheet, column_letter(fieldnames, "severity"), SEVERITY_OPTIONS, queue_sheet.max_row)
    add_validation(queue_sheet, column_letter(fieldnames, "substance_role"), SUBSTANCE_ROLE_OPTIONS, queue_sheet.max_row)
    add_validation(queue_sheet, column_letter(fieldnames, "legal_outcome"), LEGAL_OUTCOME_OPTIONS, queue_sheet.max_row)
    add_validation(queue_sheet, column_letter(fieldnames, "psychiatric_context"), PSYCHIATRIC_CONTEXT_OPTIONS, queue_sheet.max_row)

    summary_sheet = workbook.create_sheet("Summary")
    summary_fields = ["analysis_domain", "broad_cohort", "validation_priority", "n_validation_rows"]
    summary_sheet.append(summary_fields)
    for row in summary:
        summary_sheet.append([row[field] for field in summary_fields])
    style_table(summary_sheet)

    lists_sheet = workbook.create_sheet("Coding lists")
    lists = [
        ("event_present", EVENT_PRESENT_OPTIONS),
        ("false_positive_type", FALSE_POSITIVE_OPTIONS),
        ("severity", SEVERITY_OPTIONS),
        ("substance_role", SUBSTANCE_ROLE_OPTIONS),
        ("legal_outcome", LEGAL_OUTCOME_OPTIONS),
        ("psychiatric_context", PSYCHIATRIC_CONTEXT_OPTIONS),
    ]
    for col_index, (name, options) in enumerate(lists, start=1):
        lists_sheet.cell(1, col_index).value = name
        for row_index, option in enumerate(options, start=2):
            lists_sheet.cell(row_index, col_index).value = option
    style_table(lists_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def column_letter(fieldnames: list[str], field: str) -> str:
    return get_column_letter(fieldnames.index(field) + 1)


def validation_fieldnames() -> list[str]:
    return [
        "validation_priority",
        "analysis_domain",
        "broad_cohort",
        "report_id",
        "url",
        "title",
        "substance_categories",
        "erowid_categories",
        "matched_terms",
        "n_matched_terms_in_category",
        "evidence_snippets",
        "selection_reason",
        "event_present",
        "false_positive_type",
        "severity",
        "substance_role",
        "legal_outcome",
        "psychiatric_context",
        "reviewer",
        "review_notes",
    ]


def style_readme(sheet) -> None:
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 120
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")


def style_table(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "validation_priority": 16,
        "analysis_domain": 24,
        "broad_cohort": 34,
        "report_id": 14,
        "url": 42,
        "title": 42,
        "substance_categories": 34,
        "erowid_categories": 24,
        "matched_terms": 40,
        "n_matched_terms_in_category": 12,
        "evidence_snippets": 95,
        "selection_reason": 34,
        "event_present": 16,
        "false_positive_type": 26,
        "severity": 20,
        "substance_role": 18,
        "legal_outcome": 22,
        "psychiatric_context": 28,
        "reviewer": 18,
        "review_notes": 54,
    }
    headers = [cell.value for cell in sheet[1]]
    for col_index, header in enumerate(headers, start=1):
        letter = get_column_letter(col_index)
        sheet.column_dimensions[letter].width = widths.get(str(header), 22)
        cell = sheet.cell(1, col_index)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    fill_by_priority = {
        "critical": "F4CCCC",
        "high": "FCE5CD",
        "medium": "FFF2CC",
        "low": "D9EAD3",
    }
    priority_col = headers.index("validation_priority") + 1 if "validation_priority" in headers else None
    for row in sheet.iter_rows(min_row=2):
        if priority_col:
            priority = row[priority_col - 1].value
            if priority in fill_by_priority:
                row[priority_col - 1].fill = PatternFill("solid", fgColor=fill_by_priority[priority])
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a report-level validation workbook for broad AE/legal categories.")
    parser.add_argument("--reports", type=Path, default=DEFAULT_REPORTS)
    parser.add_argument("--flags", type=Path, default=DEFAULT_FLAGS)
    parser.add_argument("--out-csv", type=Path, default=DEFAULT_OUT_CSV)
    parser.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT_XLSX)
    parser.add_argument("--summary-out", type=Path, default=DEFAULT_SUMMARY)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    queue_rows = build_queue(args.reports, args.flags)
    summary = summary_rows(queue_rows)
    write_csv(args.out_csv, validation_fieldnames(), queue_rows)
    write_csv(args.summary_out, ["analysis_domain", "broad_cohort", "validation_priority", "n_validation_rows"], summary)
    build_workbook(args.out_xlsx, queue_rows, summary)

    print(f"Validation rows: {len(queue_rows)}")
    print(f"Categories sampled: {len(summary)}")
    print(f"CSV written to: {args.out_csv}")
    print(f"Workbook written to: {args.out_xlsx}")
    print(f"Summary written to: {args.summary_out}")
    for row in summary[:12]:
        print(f"  {row['validation_priority']} | {row['broad_cohort']}: {row['n_validation_rows']}")


if __name__ == "__main__":
    main()
